"""
Tests for validation export functionality.
"""

import io
import os

import pytest
from django.urls import reverse
from openpyxl import load_workbook

from validations.fake_data import CounterAPIValidationFactory, ValidationFactory


@pytest.mark.django_db
class TestValidationExport:
    def test_export_contains_reportinfo_data(self, client_authenticated_user, normal_user):
        """Test that exported Excel file contains correct reportinfo data."""
        # Create a validation with reportinfo data
        validation = ValidationFactory.create(
            core__user=normal_user,
            result_data={
                "header": {
                    "report": {"A1": "Report_Name", "B1": "Test Report"},
                    "result": ["Validation completed successfully"],
                },
                "reportinfo": {
                    "report_id": "TR",
                    "format": "tabular",
                    "cop_version": "5",
                    "institution_name": "Test University",
                    "created": "2023-01-15T10:30:00Z",
                    "created_by": "Test User",
                    "begin_date": "2023-01-01",
                    "end_date": "2023-01-31",
                },
            },
        )

        # Export the validation
        response = client_authenticated_user.get(reverse("validation-export", args=[validation.pk]))
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert str(validation.pk) in response["Content-Disposition"]
        base = os.path.splitext(validation.filename)[0]
        assert response["Content-Disposition"].startswith(f"attachment; filename={base}")
        assert response["Content-Disposition"].endswith(".xlsx")

        # Read the Excel file
        workbook = load_workbook(io.BytesIO(response.content))
        metadata_sheet = workbook["metadata"]

        # Find the rows with reportinfo data
        begin_date_value = None
        end_date_value = None
        institution_name_value = None
        created_by_value = None
        created_value = None

        for row in metadata_sheet.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] == "Begin date":
                    begin_date_value = row[1]
                elif row[0] == "End date":
                    end_date_value = row[1]
                elif row[0] == "Institution name":
                    institution_name_value = row[1]
                elif row[0] == "Created by":
                    created_by_value = row[1]
                elif row[0] == "Created":
                    created_value = row[1]

        # Verify the reportinfo data is correctly exported
        assert begin_date_value == "2023-01-01"
        assert end_date_value == "2023-01-31"
        assert institution_name_value == "Test University"
        assert created_by_value == "Test User"
        assert created_value == "2023-01-15T10:30:00Z"

    def test_export_with_missing_reportinfo_data(self, client_authenticated_user, normal_user):
        """Test that export handles missing reportinfo data gracefully."""
        # Create a validation without reportinfo data
        validation = ValidationFactory.create(
            core__user=normal_user,
            result_data={
                "header": {
                    "report": {"A1": "Report_Name", "B1": "Test Report"},
                    "result": ["Validation completed successfully"],
                }
                # No reportinfo section
            },
        )

        # Export the validation
        response = client_authenticated_user.get(reverse("validation-export", args=[validation.pk]))
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Read the Excel file
        workbook = load_workbook(io.BytesIO(response.content))
        metadata_sheet = workbook["metadata"]

        # Find the rows with reportinfo data
        begin_date_value = None
        end_date_value = None
        institution_name_value = None
        created_by_value = None
        created_value = None

        for row in metadata_sheet.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] == "Begin date":
                    begin_date_value = row[1]
                elif row[0] == "End date":
                    end_date_value = row[1]
                elif row[0] == "Institution name":
                    institution_name_value = row[1]
                elif row[0] == "Created by":
                    created_by_value = row[1]
                elif row[0] == "Created":
                    created_value = row[1]

        # Verify missing data shows as "-"
        assert begin_date_value == "-"
        assert end_date_value == "-"
        assert institution_name_value == "-"
        assert created_by_value == "-"
        assert created_value == "-"

    def test_export_counter_api_validation(self, client_authenticated_user, normal_user):
        """Test export for COUNTER API validation with reportinfo data."""
        # Create a COUNTER API validation with reportinfo data
        validation = CounterAPIValidationFactory.create(
            core__user=normal_user,
            result_data={
                "header": {
                    "report": {"A1": "Report_Name", "B1": "API Test Report"},
                    "result": ["API validation completed"],
                },
                "reportinfo": {
                    "report_id": "DR",
                    "format": "json",
                    "cop_version": "5.1",
                    "institution_name": "API University",
                    "created": "2023-02-15T14:45:00Z",
                    "created_by": "API User",
                    "begin_date": "2023-02-01",
                    "end_date": "2023-02-28",
                },
            },
        )

        # Export the validation
        response = client_authenticated_user.get(reverse("validation-export", args=[validation.pk]))
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Read the Excel file
        workbook = load_workbook(io.BytesIO(response.content))
        metadata_sheet = workbook["metadata"]

        # Find the rows with reportinfo data
        begin_date_value = None
        end_date_value = None
        institution_name_value = None
        created_by_value = None
        created_value = None

        for row in metadata_sheet.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] == "Begin date":
                    begin_date_value = row[1]
                elif row[0] == "End date":
                    end_date_value = row[1]
                elif row[0] == "Institution name":
                    institution_name_value = row[1]
                elif row[0] == "Created by":
                    created_by_value = row[1]
                elif row[0] == "Created":
                    created_value = row[1]

        # Verify the reportinfo data is correctly exported
        assert begin_date_value == "2023-02-01"
        assert end_date_value == "2023-02-28"
        assert institution_name_value == "API University"
        assert created_by_value == "API User"
        assert created_value == "2023-02-15T14:45:00Z"

    def test_export_with_partial_reportinfo_data(self, client_authenticated_user, normal_user):
        """Test that export handles partial reportinfo data correctly."""
        # Create a validation with only some reportinfo fields
        validation = ValidationFactory.create(
            core__user=normal_user,
            result_data={
                "header": {
                    "report": {"A1": "Report_Name", "B1": "Partial Test Report"},
                    "result": ["Validation completed"],
                },
                "reportinfo": {
                    "report_id": "TR",
                    "cop_version": "5",
                    "institution_name": "Partial University",
                    # Missing: created, created_by, begin_date, end_date
                },
            },
        )

        # Export the validation
        response = client_authenticated_user.get(reverse("validation-export", args=[validation.pk]))
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Read the Excel file
        workbook = load_workbook(io.BytesIO(response.content))
        metadata_sheet = workbook["metadata"]

        # Find the rows with reportinfo data
        begin_date_value = None
        end_date_value = None
        institution_name_value = None
        created_by_value = None
        created_value = None

        for row in metadata_sheet.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] == "Begin date":
                    begin_date_value = row[1]
                elif row[0] == "End date":
                    end_date_value = row[1]
                elif row[0] == "Institution name":
                    institution_name_value = row[1]
                elif row[0] == "Created by":
                    created_by_value = row[1]
                elif row[0] == "Created":
                    created_value = row[1]

        # Verify present data is exported and missing data shows as "-"
        assert begin_date_value == "-"
        assert end_date_value == "-"
        assert institution_name_value == "Partial University"
        assert created_by_value == "-"
        assert created_value == "-"

    def test_export_uses_reportinfo_instead_of_header(self, client_authenticated_user, normal_user):
        """Test that export correctly uses reportinfo data instead of header data."""
        # Create a validation with both header and reportinfo data that differ
        validation = ValidationFactory.create(
            core__user=normal_user,
            result_data={
                "header": {
                    "report": {"A1": "Report_Name", "B1": "Test Report"},
                    "result": ["Validation completed successfully"],
                    # These should NOT be used by the export
                    "begin_date": "WRONG_DATE",
                    "end_date": "WRONG_DATE",
                    "institution_name": "WRONG_INSTITUTION",
                    "created_by": "WRONG_USER",
                    "created": "WRONG_CREATED",
                },
                "reportinfo": {
                    "report_id": "TR",
                    "format": "tabular",
                    "cop_version": "5",
                    "institution_name": "Correct University",
                    "created": "2023-01-15T10:30:00Z",
                    "created_by": "Correct User",
                    "begin_date": "2023-01-01",
                    "end_date": "2023-01-31",
                },
            },
        )

        # Export the validation
        response = client_authenticated_user.get(reverse("validation-export", args=[validation.pk]))
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Read the Excel file
        workbook = load_workbook(io.BytesIO(response.content))
        metadata_sheet = workbook["metadata"]

        # Find the rows with reportinfo data
        begin_date_value = None
        end_date_value = None
        institution_name_value = None
        created_by_value = None
        created_value = None

        for row in metadata_sheet.iter_rows(values_only=True):
            if len(row) >= 2:
                if row[0] == "Begin date":
                    begin_date_value = row[1]
                elif row[0] == "End date":
                    end_date_value = row[1]
                elif row[0] == "Institution name":
                    institution_name_value = row[1]
                elif row[0] == "Created by":
                    created_by_value = row[1]
                elif row[0] == "Created":
                    created_value = row[1]

        # Verify that reportinfo data is used, not header data
        assert begin_date_value == "2023-01-01"  # From reportinfo, not "WRONG_DATE"
        assert end_date_value == "2023-01-31"  # From reportinfo, not "WRONG_DATE"
        assert (
            institution_name_value == "Correct University"
        )  # From reportinfo, not "WRONG_INSTITUTION"
        assert created_by_value == "Correct User"  # From reportinfo, not "WRONG_USER"
        assert created_value == "2023-01-15T10:30:00Z"  # From reportinfo, not "WRONG_CREATED"
